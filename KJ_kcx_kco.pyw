#! /usr/bin/env python3
# -*- Encoding: utf-8 -*-

import sqlite3
from tkinter import *
import tkinter.ttk as ttk
import datetime
from tkinter import messagebox
import os
import sys
import PyPDF2
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph
#from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill



# git
# utworzenie połączenia z bazą przechowywaną na dysku
# lub w pamięci (':memory:')
con = sqlite3.connect('baza_kcx_kco.db')

# dostęp do kolumn przez indeksy i przez nazwy
con.row_factory = sqlite3.Row

# utworzenie obiektu kursora
cur = con.cursor()





# lista zagnieżdzona przekazywania stanow kontrolek
l=200     #parametry listy-ilosc zagniezdzen
x=200    #parametry listy- ilosc zmiennych(str) w zagniezdzeniu
State=[]
for s in range(l):
    sub_list=[""]
    for y in range(x):

        sub_list.append(int(y))

        State.append(sub_list)

n=-1  #zmienna odpow. za poruszanie po bazie



# Klasa okna glownego
class Application(Frame):


    def __init__(self,master,State,n):
        """Inicjuję ramkę"""

        dt = datetime.datetime.now()
        self.now_d = dt.strftime("%Y-%m-%d")
        self.now_h = dt.strftime("%H:%M")

        super(Application,self).__init__(master)
        self.grid()
        self.nr_fab(State)                               #0
        self.kod_prod(State)                             #1
        self.nr_zlec(State)                              #2
        self.identy_nr_fab(State)                        #3
        self.tab_znam_inne_nakl(State)                   #4
        self.tab_kla_energ_ozn_kroc(State)               #5
        self.zasleki_raczki(State)                       #6
        self.estet_sil_rysy(State)                       #7
        self.kapturki_srub_zn_uziem(State)               #8
        self.izolacja_uszczelki(State)                   #9
        self.filtry(State)                              #10
        self.taca_ociekowa(State)                       #11
        self.bypass_montaz_dzialanie(State)             #12
        self.wentylatory_montaz_dzialanie(State)        #13
        self.czujniki_temper_montaz_dzial(State)        #14
        self.montaz_rozdzielnicy (State)                #15
        self.nagrz_elektr_mont_dział_alarm(State)       #16
        self.drzwi_zamki_mont_oznacz(State)             #17
        self.wydatek(State)                             #18
        self.nieszczelnosc(State)                       #29
        self.oslona_rozdzielnicy(State)                 #20
        self.dtr_ka_karta_prod_kpl_dost(State)          #21
        self.oznaczenie_KJ(State)                       #22
        self.uwagi(State)                               #23
        self.okno_uwagi(State)                          #24
        self.podpis_kontrolera(State)                   #26
        self.btn_akcept()                               #26
        self.btn_drukuj()
        self.protocol_perm()
        self.n = n
        self.btn_date()
        self.btn_zrzut()
        self.equip_protocol=str()
        self.equip_protocol_1=str()
        self.equip_protocol_2=str()



########################################################################################################################################################################################

    # poziom lini nr fabryczny


    def nr_fab(self,State):
        """numer fabryczny"""

        # utworz etykiete z zapytniem o nr fabryczny
        self.lbl_dist_0=Label(self)
        self.lbl_dist_0.grid(row = 0, column = 0 , padx=6)                       #dystans col 0

        self.lbl_nr_fab = Label(self, text ="Podaj nr fabryczny")
        self.lbl_nr_fab.grid(row = 0, column = 1 , sticky = W )

        self.lbl_dist_2=Label(self)
        self.lbl_dist_2.grid(row = 0, column = 2 , padx=5)                               #dystans  col  2

        self.lbl_dist_3=Label(self)
        self.lbl_dist_3.grid(row = 0, column = 3 , padx=1)                               #dystans   col 3



        # utworz widzet Entry do przyjecia nr_fab

        var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1]==0 :
            var1.set("")
        else:
            var1.set(State[0][1])


        self.ent_nr_fab= Entry(self,textvariable=var1)
        self.ent_nr_fab.grid(row=0, column = 4, ipadx=10 )

        self.lbl_dist_5=Label(self)
        self.lbl_dist_5.grid(row = 0, column = 5 , padx=4)                               #dystans col  5



        #utworz przyciski 'archiwum'5

        self.btn_up = Button(self,text= "UP", command=self.arch_UP )
        self.btn_up.grid(row = 0, column=6 , ipadx=20 , sticky=E,padx=(15,0) )

        self.btn_dn = Button(self,text= "DN", command=self.arch_DOWN )
        self.btn_dn.grid(row = 0, column=7 ,ipadx=20, sticky=W)




        self.lbl_dist_8=Label(self)
        self.lbl_dist_8.grid(row = 0, column = 5 , padx=4, pady=3)                               #dystans col  8


########################################################################################################################################################################################



    def kod_prod(self,State):



        self.czynnosc = Label(self, text ="Podaj kod produktu")
        self.czynnosc.grid(row = 1, column = 1 , sticky = W )


        self.typ_ahu = StringVar()
        self.ent_kod_prod = ttk.Combobox(self, textvariable = self.typ_ahu )
        self.ent_kod_prod.bind("<<ComboboxSelected>>", self.onReturn)
        self.ent_kod_prod.grid(row = 1, column = 4,   sticky =W)
        self.ent_kod_prod['values'] = ('','KCX-300', 'KCX-500','KCX-800','KCX-1200','KCO-300','KCO-500','KCO-800','KCO-1200','KCX 300 EPP','KCX 500 EPP','KCX 800 EPP',
                                       'KCO 300 EPP','KCO 500 EPP','KCO 800 EPP')
        self.ent_kod_prod.current(0)
        a=State[0][2]

        if a=='KCX-300':
            self.ent_kod_prod.current(1)
        if a=='KCO-300':
            self.ent_kod_prod.current(5)


        if a=='KCX-500' :
            self.ent_kod_prod.current(2)
        if  a=='KCO-500':
            self.ent_kod_prod.current(6)


        if a=='KCX-800' :
            self.ent_kod_prod.current(3)
        if a == 'KCO-800':
            self.ent_kod_prod.current(7)


        if a=='KCX-1200':
            self.ent_kod_prod.current(4)
        if a == 'KCO-1200':
            self.ent_kod_prod.current(8)

        if a == 'KCX 300 EPP':
            self.ent_kod_prod.current(9)
        if a == 'KCO 300 EPP':
            self.ent_kod_prod.current(12)

        if a == 'KCX 500 EPP':
            self.ent_kod_prod.current(10)
        if a == 'KCO 500 EPP':
            self.ent_kod_prod.current(13)

        if a == 'KCX 800 EPP':
            self.ent_kod_prod.current(11)
        if a == 'KCO 800 EPP':
            self.ent_kod_prod.current(14)



        var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1]==0 :
            var1.set("")
            Time_label_txt=""
        else:
            var1.set(State[0][2])
            Time_label_txt ="Ostatnie badanie:"





         # Wyswietlanie komunikatu
        self.lbl_t1 = Label(self, text=Time_label_txt)
        self.lbl_t1.grid(row = 1, column =6, columnspan=2)





###################################################################################################################################################################################


    def onReturn(self, event):  # modół sterujący wyświetlaniem widzetów- Event
        self.prod_cod = str(self.typ_ahu.get())
        self.podpowiedz()
        self.ent_wydatek.delete(0, END)
        self.ent_nieszczel.delete(0, END)

        if str(self.typ_ahu.get()) in ('KCX 300 EPP','KCX 500 EPP','KCX 800 EPP','KCO 300 EPP','KCO 500 EPP','KCO 800 EPP' ):

            self.ent_wydatek.insert(END,"Nie wymagane")
            self.ent_nieszczel.insert(END,"Nie wymagane")





#########################################################################################################################################################################################

    def podpowiedz(self):    # modół sterujący wyświetlaniem widzetów- Event

        if self.prod_cod=='KCX-1200'or self.prod_cod=='KCO-1200' :
             messagebox.showinfo('info','Wybrano model 1200 bez nagrzewnicy wbudowanej')

             self.lbl_mont_dzial = Label(self, text=(State[0][17]))
             self.lbl_mont_dzial.configure(text='Brak')
             self.lbl_mont_dzial.grid(row=16, column=7)
             State[0][17] = 'Brak'
             self.nagrz_elektr_mont_dział_alarm(State)

        if self.prod_cod=='KCO-300'or self.prod_cod=='KCO-500' or self.prod_cod=='KCO-800' or self.prod_cod=='KCO-1200' :
            messagebox.showinfo('info', ' Wybrano model z wymiennikiem obrotowym')
            self.lbl_bypas_wym_obr.configure(text='Wymiennik obrotowy')



        if self.prod_cod=='KCX-300'or self.prod_cod=='KCX-500' or self.prod_cod=='KCX-800' or self.prod_cod=='KCX-1200' :
            messagebox.showinfo('info', ' Wybrano model z wymiennikiem przeciwprądowym')
            self.lbl_bypas_wym_obr.configure(text='Bypass - montaż i działanie, szczelność')

        if self.prod_cod == 'KCO 300 EPP' or self.prod_cod == 'KCO 500 EPP' or self.prod_cod == 'KCO 800 EPP' :
            messagebox.showinfo('info', ' Wybrano model z wymiennikiem obrotowym')
            self.lbl_bypas_wym_obr.configure(text='Wymiennik obrotowy')

        if self.prod_cod == 'KCX 300 EPP' or self.prod_cod == 'KCX 500 EPP' or self.prod_cod == 'KCX 800 EPP' :
            messagebox.showinfo('info', ' Wybrano model z wymiennikiem przeciwprądowym')
            self.lbl_bypas_wym_obr.configure(text='Bypass - montaż i działanie, szczelność')



    ########################################################################################################################################################################################


    def nr_zlec(self,State):

        # utworz etykiete z zapytniem o zlecenie
        self.czynnosc = Label(self, text ="Podaj nr zlecenia")
        self.czynnosc.grid(row = 2, column = 1 , columnspan = 1, sticky = W )






        # utworz widzet Entry do przyjecia zlecenia


        var2=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera
        Time_txt=StringVar()
        t1=(str(State[0][27])+' '+str(State[0][28]))

        if State[0][1]==0 :
            var2.set("")
            t1=str(self.now_d+'  '+str(self.now_h))
            Time_txt.set(t1)

        else:
            var2.set(State[0][3])
            t1=(str(State[0][27])+' '+str(State[0][28]))
            Time_txt.set(t1)
            self.now_d = str(State[0][27])
            self.now_h = str(State[0][28])


        self.ent_nr_zlec= Entry(self,textvariable=var2)
        self.ent_nr_zlec.grid(row=2, column = 4, ipadx=10)

        self.lbl_dist_10=Label(self)
        self.lbl_dist_10.grid(row = 2, column = 5 , pady=3)




        self.lbl_t2 = Entry(self, text = Time_txt, state = DISABLED)
        self.lbl_t2.grid(row = 2, column =6, columnspan=2 )

        self.lbl_dist_10=Label(self)
        self.lbl_dist_10.grid(row = 28, column = 5 , pady=3)
        var3=StringVar()
        a=var3.set("wprowadz datę")
        self.ent_data= Entry(self, textvariable=var3 )
        self.ent_data.grid(row=29, column = 1, ipadx=15)

        #contens3 = str(self.ent_nr_zlec.get())




    def btn_date(self):
        self.nowa_data = Button(self, text ="Nowa data",command = self.give_new_date)
        self.nowa_data.grid(row = 29, column = 3)

    def give_new_date(self):
        self.now_d = self.ent_data.get()





########################################################################################################################################################################################


    def identy_nr_fab(self,State):


        self.czynnosc = Label(self, text ="Identyfikacja , nalepka z nr. fabrycznym ")
        self.czynnosc.grid(row = 3, column = 1,sticky = W )
        self.ident_tab_znam = StringVar()
        self.ident_tab_znam.set(State[0][4])



        Radiobutton(self,
                    text =  "Tak",
                    variable = self.ident_tab_znam,
                    value = "Pozytyw",
                    ).grid(row = 3, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.ident_tab_znam,
                    value = "Negatyw",
                    ).grid(row = 3, column = 4, sticky=E)



    # Wyswietlanie stanu   z bazy
        if State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][4] )
            self.lbl_czynnosc_info.grid(row = 3, column = 7)

#######################################################################################################################################################################################



    def tab_znam_inne_nakl(self,State):

        self.czynnosc = Label(self, text ="Tabl. znamionowa, inne naklejki")
        self.czynnosc.grid(row = 4, column = 1,sticky = W )

        self.tab_znam_nakl = StringVar()

        self.tab_znam_nakl.set(State[0][5])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.tab_znam_nakl,
                    value = "Pozytyw",
                    ).grid(row = 4, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.tab_znam_nakl,
                    value = "Negatyw",
                   #command = self.update_text
                    ).grid(row = 4, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][5])
            self.lbl_czynnosc_info.grid(row = 4, column = 7)




#######################################################################################################################################################################################


    def tab_kla_energ_ozn_kroc(self,State):

        self.czynnosc = Label(self, text ="Tab.klasa energetyczna, ozn. krócców")
        self.czynnosc.grid(row = 5, column = 1,sticky = W )

        self.kla_en_oz_kr = StringVar()

        self.kla_en_oz_kr.set(State[0][6])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.kla_en_oz_kr,
                    value = "Pozytyw",
                   # command = self.update_text
                    ).grid(row = 5, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.kla_en_oz_kr,
                    value = "Negatyw",
                   #command = self.update_text
                    ).grid(row = 5, column = 4,sticky=E)


       # Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][6])
            self.lbl_czynnosc_info.grid(row = 5, column = 7)






######################################################################################################################################################################################




    def zasleki_raczki(self,State):

        self.lbl_czynnosc = Label(self, text ="Zaślepki srub, rączki")
        self.lbl_czynnosc.grid(row = 6, column = 1,sticky = W )

        self.zasl_raczki= StringVar()

        self.zasl_raczki.set(State[0][7])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.zasl_raczki,
                    value = "Pozytyw",
                    ).grid(row = 6, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.zasl_raczki,
                    value = "Negatyw",
                    ).grid(row = 6, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][7])
            self.lbl_czynnosc_info.grid(row = 6, column = 7)


#######################################################################################################################################################################################


    def estet_sil_rysy(self,State):


        self.lbl_czynnosc = Label(self, text ="Estetyka - silikonowanie , porysowania ")
        self.lbl_czynnosc.grid(row = 7, column = 1,sticky = W )

        self.est_silik_rysy = StringVar()

        self.est_silik_rysy.set(State[0][8])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.est_silik_rysy ,
                    value = "Pozytyw",
                    ).grid(row = 7, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.est_silik_rysy ,
                    value = "Negatyw",
                    ).grid(row = 7, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][8])
            self.lbl_czynnosc_info.grid(row = 7, column = 7)



########################################################################################################################################################################################


    def kapturki_srub_zn_uziem(self,State):

        self.lbl_czynnosc = Label(self, text ="Kapturki srub nagrzew, znaczki uziemień" )
        self.lbl_czynnosc.grid(row = 8, column = 1,sticky = W )

        self.kapt_srub_uziem = StringVar()

        self.kapt_srub_uziem.set(State[0][9])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.kapt_srub_uziem,
                    value = "Pozytyw",
                    ).grid(row = 8, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.kapt_srub_uziem,
                    value = "Negatyw",
                    ).grid(row = 8, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][9])
            self.lbl_czynnosc_info.grid(row = 8, column = 7)









########################################################################################################################################################################################


    def izolacja_uszczelki(self,State):


        self.lbl_czynnosc = Label(self, text ="Izolacja, uszczelki" )
        self.lbl_czynnosc.grid(row = 9, column = 1,sticky = W )

        self.izol_uszcz = StringVar()

        self.izol_uszcz.set(State[0][10])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.izol_uszcz,
                    value = "Pozytyw",
                    ).grid(row = 9, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.izol_uszcz,
                    value = "Negatyw",
                    ).grid(row = 9, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][10])
            self.lbl_czynnosc_info.grid(row = 9, column = 7)


########################################################################################################################################################################################


    def filtry (self,State):


        self.lbl_czynnosc = Label(self, text ="Filtry" )
        self.lbl_czynnosc.grid(row = 10, column = 1,sticky = W )

        self.mon_filtr = StringVar()

        self.mon_filtr.set(State[0][11])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.mon_filtr,
                    value = "Pozytyw",
                    ).grid(row = 10, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.mon_filtr,
                    value = "Negatyw",
                    ).grid(row = 10, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][11])
            self.lbl_czynnosc_info.grid(row = 10, column = 7)




########################################################################################################################################################################################



    def taca_ociekowa(self,State):


        self.lbl_czynnosc = Label(self, text ="Taca ociekowa - szczel.-malow.-silk-nie")
        self.lbl_czynnosc.grid(row = 11, column = 1,sticky = W )

        self.tac_ociek = StringVar()

        self.tac_ociek.set(State[0][12])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.tac_ociek,
                    value = "Pozytyw",
                    ).grid(row = 11, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.tac_ociek,
                    value = "Negatyw",
                    ).grid(row = 11, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][12])
            self.lbl_czynnosc_info.grid(row = 11, column = 7)






##########################################################################################################################################################################################
    def bypass_montaz_dzialanie(self,State):


        self.lbl_bypas_wym_obr = Label(self, text ="Bypass - montaż i działanie, szczelność" )
        self.lbl_bypas_wym_obr.grid(row = 12, column = 1,sticky = W )

        self.mon_bypas = StringVar()

        self.mon_bypas.set(State[0][13])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.mon_bypas,
                    value = "Pozytyw",
                    ).grid(row = 12, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.mon_bypas,
                    value = "Negatyw",
                    ).grid(row = 12, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][13])
            self.lbl_czynnosc_info.grid(row = 12, column = 7)


####################################################################################################################################################################################


    def wentylatory_montaz_dzialanie(self,State):


        self.lbl_czynnosc = Label(self, text ="Wentylatory- montaż , działnie" )
        self.lbl_czynnosc.grid(row = 13, column = 1,sticky = W )

        self.went = StringVar()

        self.went.set(State[0][14])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.went,
                    value = "Pozytyw",
                    ).grid(row = 13, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.went,
                    value = "Negatyw",
                    ).grid(row = 13, column = 4,sticky=E)



# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][14])
            self.lbl_czynnosc_info.grid(row = 13, column = 7)


#######################################################################################################################################################################################

    def czujniki_temper_montaz_dzial(self,State):


        self.lbl_czynnosc = Label(self, text ="Czujniki temeratur -montaż, działanie" )
        self.lbl_czynnosc.grid(row = 14, column = 1,sticky = W )

        self.czujn_temp = StringVar()

        self.czujn_temp.set(State[0][15])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.czujn_temp,
                    value = "Pozytyw",
                    ).grid(row = 14, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.czujn_temp,
                    value = "Negatyw",
                    ).grid(row = 14, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][15])
            self.lbl_czynnosc_info.grid(row = 14, column = 7)


#######################################################################################################################################################################################






    def montaz_rozdzielnicy (self,State):

        self.lbl_czynnosc = Label(self, text ="Montaż rozdzielnicy")
        self.lbl_czynnosc.grid(row = 15, column = 1,sticky = W )

        self.rozdz_mont = StringVar()

        self.rozdz_mont.set(State[0][14])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.rozdz_mont,
                    value = "Pozytyw",
                    ).grid(row = 15, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.rozdz_mont,
                    value = "Negatyw",
                    ).grid(row = 15, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][16])
            self.lbl_czynnosc_info.grid(row = 15, column = 7)


#######################################################################################################################################################################################


    def nagrz_elektr_mont_dział_alarm (self,State):


        self.lbl_czynnosc = Label(self, text ="Nagrzewnica- montaż, działenie, alarm " )
        self.lbl_czynnosc.grid(row = 16, column = 1,sticky = W )

        self.nagrz_dzial = StringVar()

        self.nagrz_dzial.set(State[0][17])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.nagrz_dzial,
                    value = "Pozytyw",
                    ).grid(row = 16, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.nagrz_dzial,
                    value = "Negatyw",
                    ).grid(row = 16, column = 4,sticky=E)


        Radiobutton(self,
                    text =  "Brak",
                    variable = self.nagrz_dzial,
                    value = "Brak",
                    ).grid(row = 16, column = 6,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_mont_dzial = Label(self,text=State[0][17])
            self.lbl_mont_dzial.grid(row = 16, column = 7)


#######################################################################################################################################################################################


    def drzwi_zamki_mont_oznacz (self,State):

        self.lbl_czynnosc = Label(self, text ="Drzwi, zamki, montaz, oznaczenia, rączki" )
        self.lbl_czynnosc.grid(row = 17, column = 1,sticky = W )

        self.drzwi = StringVar()

        self.drzwi.set(State[0][18])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.drzwi,
                    value = "Pozytyw",
                    ).grid(row = 17, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.drzwi,
                    value = "Negatyw",
                    ).grid(row = 17, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][18])
            self.lbl_czynnosc_info.grid(row = 17, column = 7)



######################################################################################################################################################################################


    def wydatek(self,State):

        # utworz etykiete z zapytniem o wydatek
        self.lbl_dist_0=Label(self)
        self.lbl_dist_0.grid(row = 18, column = 0 , padx=6)                       #dystans col 0

        self.lbl_nr_fab = Label(self, text ="Podaj wydatek [m/s]")
        self.lbl_nr_fab.grid(row = 18, column = 1 , sticky = W )

        self.lbl_dist_2=Label(self)
        self.lbl_dist_2.grid(row = 18, column = 2 , padx=5)                               #dystans  col  2

        self.lbl_dist_3=Label(self)
        self.lbl_dist_3.grid(row = 18, column = 3 , padx=1)                               #dystans   col 3



        # utworz widzet Entry do przyjecia nr_fab

        var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1]==0 :
            var1.set("")
        else:
            var1.set(State[0][19])

        if State[0][19] == -1:
            var1.set("Nie dotyczy")



        self.ent_wydatek= Entry(self,textvariable=var1)
        self.ent_wydatek.grid(row=18, column = 4)


#######################################################################################################################################################################################

    def nieszczelnosc(self,State):

        # utworz etykiete z zapytniem o nr nieszczelnosc
        self.lbl_dist_0=Label(self)
        self.lbl_dist_0.grid(row = 19, column = 0 , padx=6)                              #dystans col 0

        self.lbl_nr_fab = Label(self, text ="Podaj nieszczelność [m/s]")
        self.lbl_nr_fab.grid(row = 19, column = 1 , sticky = W )

        self.lbl_dist_2=Label(self)
        self.lbl_dist_2.grid(row = 19, column = 2 , padx=5)                               #dystans  col  2

        self.lbl_dist_3=Label(self)
        self.lbl_dist_3.grid(row = 19, column = 3 , padx=1)                               #dystans   col 3



        # utworz widzet Entry do przyjecia niesczelnosci

        var1=StringVar()    # zmienna pomocnicza - ukrywanie wywswl - zera

        if State[0][1]==0 :
            var1.set("")
        else:
            var1.set(State[0][20])


        if State[0][20] == -1 :
            var1.set("Nie dotyczy")



        self.ent_nieszczel= Entry(self,textvariable=var1)
        self.ent_nieszczel.grid(row=19, column = 4)



###############################################################################################################################################################################


    def oslona_rozdzielnicy  (self,State):


        self.lbl_czynnosc = Label(self, text ="Osłona rozdzielnicy z zn. ostrzgawczym" )
        self.lbl_czynnosc.grid(row = 20, column = 1,sticky = W )

        self.osl_rozdz = StringVar()

        self.osl_rozdz.set(State[0][21])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.osl_rozdz,
                    value = "Pozytyw",
                    ).grid(row = 20, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.osl_rozdz,
                    value = "Negatyw",
                    ).grid(row = 20, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][21])
            self.lbl_czynnosc_info.grid(row = 20, column = 7)






#######################################################################################################################################################################################


    def dtr_ka_karta_prod_kpl_dost (self,State):


        self.lbl_czynnosc = Label(self, text ="DTR-ka, karta prod-u, kompl. dostawy" )
        self.lbl_czynnosc.grid(row = 21, column = 1,sticky = W )

        self.dtr_kart_prod = StringVar()

        self.dtr_kart_prod.set(State[0][22])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.dtr_kart_prod,
                    value = "Pozytyw",
                    ).grid(row = 21, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.dtr_kart_prod,
                    value = "Negatyw",
                    ).grid(row = 21, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][22])
            self.lbl_czynnosc_info.grid(row = 21, column = 7)




#######################################################################################################################################################################################


    def oznaczenie_KJ (self,State):


        self.lbl_czynnosc = Label(self, text ="Oznaczene KJ" )
        self.lbl_czynnosc.grid(row = 22, column = 1,sticky = W )

        self.ozn_KJ = StringVar()

        self.ozn_KJ.set(State[0][23])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.ozn_KJ,
                    value = "Pozytyw",
                    ).grid(row = 22, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.ozn_KJ,
                    value = "Negatyw",
                    ).grid(row = 22, column = 4,sticky=E)


# Wyswietlanie stanu  z bazy

        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][23])
            self.lbl_czynnosc_info.grid(row = 22, column = 7)




#######################################################################################################################################################################################

    def uwagi(self,State):

        self.lbl_czynnosc = Label(self,text ="Uwagi")
        self.lbl_czynnosc.grid(row = 23, column = 1 , sticky =W)  # dystans


        self.uwagi_bool = StringVar()
        self.uwagi_bool.set(State[0][24])


        Radiobutton(self,
                    text =  "Tak",
                    variable = self.uwagi_bool,
                    value = "Pozytyw",
                    ).grid(row = 23, column = 4,sticky=W)

        Radiobutton(self,
                    text =  "Nie",
                    variable = self.uwagi_bool,
                    value = "Negatyw",
                    ).grid(row = 23, column = 4,sticky=E)

# Wyswietlanie stanu  z bazy


        if  State[0][1]!=0 :
            self.lbl_czynnosc_info = Label(self,text=State[0][24])
            self.lbl_czynnosc_info.grid(row = 23, column = 7)


#######################################################################################################################################################################################


    def okno_uwagi(self,State):

        var3=StringVar() # zmienna pomocnicza - ukrywanie wywswl - zera
        var3=State [0][25]
        self.uwagi_txt = Text(self, width = 65, height = 1, wrap=CHAR )
        self.uwagi_txt.grid(row = 24, column = 1, columnspan = 10 , sticky =W)

        if State[0][1]==0 :
            self.uwagi_txt.insert(END," ")
        else:
            self.uwagi_txt.insert(END,var3)

####################################################################################################################################################################self.equip_protocol#############

#dystans

        self.lbl_dist_10=Label(self)
        self.lbl_dist_10.grid(row = 25, column = 1)


####################################################################################################################################################################################

    def podpis_kontrolera(self,State):

        self.podpis = StringVar()
        self.combobox = ttk.Combobox(self, textvariable = self.podpis)
        self.combobox.grid(row = 26, column = 3, columnspan = 2 , sticky =W)
        self.combobox['values'] = ('','Jarosław Olszewski', 'Piotr Tylak','Dominik Tanski','Agnieszka Karnas','Marcin Pisarski','Monika Roman','Tomasz Makowski' )
        self.combobox.current(0)
        a=State[0][26]
        if a=='Jaroslaw Olszewski' or a=='Jarosław Olszewski':
            self.combobox.current(1)
        if a=='Piotr Tylak':
            self.combobox.current(2)

        if a=='Dominik Tański'or a=='Dominik Tanski':
            self.combobox.current(3)

        if a=='Agnieszka Karnas':
            self.combobox.current(4)

        if a=='Marcin Pisarski':
            self.combobox.current(5)

        if a=='Monika Roman':
            self.combobox.current(6)

        if a=='Tomasz Makowski':
            self.combobox.current(7)

############################################################################################################################################################################




######################################################################################################################################################################################


    #poziom lini przycisk akceptuj

    def btn_akcept(self):
        #utworz przycisk - akceptuj- poziom
        self.submit_bttn = Button(self, text ="Akceptuj", command = self.ostrzezenie_zapis)
        self.submit_bttn.grid(row = 26, column = 1)


    def ostrzezenie_zapis(self):
        if messagebox.askyesno("Zapis danych", "Czy zapisać dane ?"):
            self.equipment()
            self.akcept()


    def btn_drukuj(self):
        self.print_bttn = Button(self, text ="Drukuj",command = self.info_druk)
        self.print_bttn.grid(row = 26, column = 7)



    def info_druk(self):

        if messagebox.askyesno("Wydruk do PDF", "Czy zapisać i wydrukować do PDF ?"):
            self.equipment()
            self.control_list(State)
            self.filling_factory(State)
            #self.akcept()


    def equipment(self):
        if self.typ_ahu.get() in ('KCX-300', 'KCX-500', 'KCX-800', 'KCX-1200','KCX 300 EPP','KCX 500 EPP','KCX 800 EPP'):
            self.equip_lista = 'Bypass-montaż-działanie'
            self.equip_protocol='Bypass'

        if self.typ_ahu.get() in ('KCO-300', 'KCO-500', 'KCO-800', 'KCO-1200', 'KCO 300 EPP', 'KCO 500 EPP', 'KCO 800 EPP'):
            self.equip_lista = 'Wym.obrotowy.-montaż-działanie'
            self.equip_protocol_1='Wymiennik'
            self.equip_protocol_2='obrotowy'

    # funkcja przycisku akeptuj
    def akcept (self):

        contens1 = str(self.ent_nr_fab.get())
        contens2 = self.typ_ahu.get()
        contens3 = str(self.ent_nr_zlec.get())
        id1 = self.ident_tab_znam.get()
        id2 = self.tab_znam_nakl.get()
        id3 = self.kla_en_oz_kr.get()
        id4 = self.zasl_raczki.get()
        id5 = self.est_silik_rysy.get()
        id6 = self.kapt_srub_uziem.get()
        id7 = self.izol_uszcz.get()
        id8 = self.mon_filtr.get()
        id9 = self.tac_ociek.get()
        id10 = self.mon_bypas.get()
        id11 = self.went.get()
        id12 = self.czujn_temp.get()
        id13 = self.rozdz_mont.get()
        id14 = self.nagrz_dzial.get()
        id15 = self.drzwi.get()
        id16 = self.ent_wydatek.get()
        id17 = self.ent_nieszczel.get()

        if self.typ_ahu.get() in ('KCX-300', 'KCX-500', 'KCX-800', 'KCX-1200', 'KCO 300', 'KCO 500', 'KCX 800'):
            try:
                float(id16.replace(',', '.'))
                float(id17.replace(',', '.'))
            except:
                self.warning()

        if self.typ_ahu.get() in ('KCX 300 EPP', 'KCX 500 EPP', 'KCX 800 EPP', 'KCX 1200 EPP', 'KCO 300 EPP', 'KCO 500 EPP', 'KCO 800 EPP'):
            id16  = '-1'
            id17 = '-1'

        id18 = self.osl_rozdz.get()
        id19 = self.dtr_kart_prod.get()
        id20 = self.ozn_KJ.get()
        id21 = self.uwagi_bool.get()
        id22 = self.uwagi_txt.get(1.0, END)
        id23 = self.podpis.get()
        id24 = str(self.now_d)
        id25 = str(self.now_h)
        cur.execute('INSERT INTO tab VALUES (NULL,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,DATE(?),?);',(contens1,contens2,contens3,\
        id1,id2,id3,id4,id5,id6,id7,id8,id9,id10,id11,id12,id13,id14,id15,id16,id17,id18,id19,id20,id21,id22,id23,id24,id25))

        messagebox.showinfo("Zapis danych", "Zapisano")
        con.commit()



    def btn_zrzut(self):
        self.print_bttn = Button(self, text ="Zrzut",command = self.zrzut)
        self.print_bttn.grid(row = 29, column = 7)


    def zrzut(self):

        fill_color='BFBFBF'#'a5a391'

        excelfile = openpyxl.load_workbook('C:\\Backup\\KCX_Lista_temp.xlsx')  # open a excel file with .xlsx format
        excelfile.get_sheet_names()  # get names of all spreadsheet in the file
        sheet1 = excelfile.get_sheet_by_name("2019")  # get the first spreadsheet by name
        nr_line = sheet1.max_row  # get the number of rows in the sheet



        uwag= sheet1.cell(row = (nr_line)+1, column = 5)

        uwag.value = self.uwagi_txt.get(1.0, END)





        uwag.alignment= Alignment(horizontal='center')

        a=self.uwagi_bool.get()

        if a=='Pozytyw':
            wypelnienie=True
        else:
            wypelnienie=False

        if wypelnienie==True:
            uwag.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")


        numfab  = sheet1.cell(row = (nr_line)+1, column = 2 )
        numfab.value = str(self.ent_nr_fab.get())
        numfab.alignment= Alignment(horizontal='center')

        if wypelnienie==True:
            numfab.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")



        empty_01 = sheet1.cell(row = (nr_line)+1, column = 3 )
        if wypelnienie==True:
            empty_01.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")


        kod_prod  = sheet1.cell(row = (nr_line)+1, column = 4 )
        kod_prod.value = self.typ_ahu.get()
        kod_prod.alignment= Alignment(horizontal='center')
        if wypelnienie==True:
            kod_prod.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")




        Year,month,day=self.now_d.split('-')
        data_format= day + '-' + month + '-'+Year
        data= sheet1.cell(row = (nr_line)+1, column = 6)
        data.value=data_format
        data.alignment= Alignment(horizontal='center')
        if wypelnienie==True:
            data.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")


        #messagebox.showinfo('fddf',data_raw)
        zlec= sheet1.cell(row = (nr_line)+1, column = 7)
        zlec.value = int(self.ent_nr_zlec.get())
        zlec.alignment= Alignment(horizontal='center')
        if wypelnienie==True:
            zlec.fill = PatternFill(start_color= fill_color, end_color= fill_color, fill_type="solid")






        try:
            #('C:\\Users\\jolszewski\Desktop\\Expenses01.xlsx')

            excelfile.save('C:\\Backup\\KCX_Lista_temp.xlsx')###
            messagebox.showinfo('info','Zrzut do Excel poprawny')
        except:

            messagebox.showinfo('info','Cos poszło nie tak')



    #  funkcja zwiększająca zmeinną sterującą bazą
    def arch_UP(self):
        self.n+=1
        self.arch(self.n)

    #  funkcja zmniejszająca zmeinną sterującą bazą
    def arch_DOWN(self):
        self.n-=1
        self.arch(self.n)


#########################################################################################################################################################################################
#  wydruk PDF

    def control_list(self,State):

      #  deklaracja zmiennych


        self.pt_contens1 = str(self.ent_nr_fab.get())
        self.pt_contens2 = self.typ_ahu.get()
        self.pt_contens3 = str(self.ent_nr_zlec.get())
        self.pt_id1 = self.ident_tab_znam.get()
        self.pt_id2 = self.tab_znam_nakl.get()
        self.pt_id3 = self.kla_en_oz_kr.get()
        self.pt_id4 = self.zasl_raczki.get()
        self.pt_id5 = self.est_silik_rysy.get()
        self.pt_id6 = self.kapt_srub_uziem.get()
        self.pt_id7 = self.izol_uszcz.get()
        self.pt_id8 = self.mon_filtr.get()
        self.pt_id9 = self.tac_ociek.get()
        self.pt_id10 = self.mon_bypas.get()
        self.pt_id11 = self.went.get()
        self.pt_id12 = self.czujn_temp.get()
        self.pt_id13 = self.rozdz_mont.get()
        self.pt_id14 = self.nagrz_dzial.get()
        self.pt_id15 = self.drzwi.get()
        self.pt_id16 = self.ent_wydatek.get()
        self.pt_id17 = self.ent_nieszczel.get()

        if self.typ_ahu.get() in ('KCX-300', 'KCX-500', 'KCX-800', 'KCX-1200', 'KCO 300', 'KCO 500', 'KCX 800'):

            try:
                float(self.pt_id16.replace(',', '.'))
                float(self.pt_id17.replace(',', '.'))
            except:
                self.warning()

        if self.typ_ahu.get() in ('KCX 300 EPP', 'KCX 500 EPP', 'KCX 800 EPP', 'KCX 1200 EPP', 'KCO 300 EPP', 'KCO 500 EPP', 'KCO 800 EPP'):
            self.pt_id16 = ''
            self.pt_id17 = ''



        self.pt_id18 = self.osl_rozdz.get()
        self.pt_id19 = self.dtr_kart_prod.get()
        self.pt_id20 = self.ozn_KJ.get()
        self.pt_id21 = self.uwagi_bool.get()
        self.pt_id22 = self.uwagi_txt.get("1.0", END)
        self.pt_id23 = self.podpis.get()
        self.pt_time = str(self.now_d+ " "+self.now_h)



        pdfmetrics.registerFont(TTFont('DejaMono', 'DejaVuSansMono.ttf'))

        doc = SimpleDocTemplate("__file__"+"../../listy_kontrolne/"+self.pt_contens1+"_lista_kontr_KCX.pdf", pagesize=A4 )

        styleSheet = getSampleStyleSheet()

        style = styleSheet['BodyText']

        elements = []



        #graniczne przecieki

        leak_bool='Negatyw'
        if self.pt_contens2=='KCX-300' and float(self.pt_id17)<0.3:
            leak_bool='Pozytyw'
        if self.pt_contens2=='KCO-300' and float(self.pt_id17)<0.3:
            leak_bool='Pozytyw'

        if self.pt_contens2=='KCX-500' and float(self.pt_id17)<0.4:
            leak_bool='Pozytyw'
        if self.pt_contens2=='KCO-500' and float(self.pt_id17)<0.4:
            leak_bool='Pozytyw'


        if self.pt_contens2=='KCX-800' and float(self.pt_id17)<0.5:
            leak_bool='Pozytyw'
        if self.pt_contens2=='KCO-800' and float(self.pt_id17)<0.5:
            leak_bool='Pozytyw'


        if self.pt_contens2=='KCX-1200' and float(self.pt_id17)<0.3:
            leak_bool='Pozytyw'
        if self.pt_contens2=='KCO-1200' and float(self.pt_id17)<0.3:
            leak_bool='Pozytyw'

        if self.pt_contens2 in ('KCX 300 EPP', 'KCX 500 EPP', 'KCX 800 EPP', 'KCX 1200 EPP', 'KCO 300 EPP', 'KCO 500 EPP', 'KCO 800 EPP'):
            leak_bool = 'Pozytyw'








        #ocena sumaryczna
        self.summary_mark ="Negatyw"
        self.protocol_permission=False

        if self.pt_id1=="Pozytyw" and self.pt_id2=="Pozytyw" and self.pt_id3=="Pozytyw" and self.pt_id4=="Pozytyw" and self.pt_id4=="Pozytyw"\
        and self.pt_id5=="Pozytyw" and self.pt_id6=="Pozytyw" and self.pt_id7=="Pozytyw" and self.pt_id7=="Pozytyw" and self.pt_id8=="Pozytyw"\
        and self.pt_id9=="Pozytyw" and self.pt_id10=="Pozytyw" and self.pt_id11=="Pozytyw" and self.pt_id12=="Pozytyw" and self.pt_id13=="Pozytyw" and leak_bool=='Pozytyw'\
        and  self.pt_id14 in ("Pozytyw","Brak")and self.pt_id15=="Pozytyw" and self.pt_id18=="Pozytyw" and self.pt_id19=="Pozytyw" and self.pt_id20=="Pozytyw"\
        and self.pt_id21=="Pozytyw" :
            self.summary_mark ="Pozytyw"
            self.protocol_permission=True

        if self.pt_contens2 in ('KCX-300', 'KCX-500', 'KCX-800', 'KCX-1200', 'KCO-300', 'KCO-500', 'KCO-800','KC0-1200'):
            wydatek_info  = 'Wydatek'
            szczelnosc_info = 'Szczelnosc'
            unit_velocity = 'm/s'
        else :
            wydatek_info = 'Nie dotyczy'
            szczelnosc_info = 'Nie dotyczy'
            unit_velocity = ''
            leak_bool = ''




        data= [['Logo','','Montaż aparatów KCX            ZP3 SKOWARCZ' ],
               ['','','Lista kontrolna czynnosci sprawdzajacych KCX'],
               ['Lp', 'Operacja', 'Lp', 'Wytyczne', 'Ocena'],
               ['1', 'Kontrola wizualna', '1.1', 'Identyfikacja, naklejka z nr. fab ',self.pt_id1],
               ['' , '', '1.2', 'Tab. znaminowa, inne nalepki', self.pt_id2],
               ['', '' , '1.3', 'Tab. klasa energetyczna, ozn króćców', self.pt_id3],
               ['', '', '1.4', 'Zaślepki srub, rączki', self.pt_id4],
               ['', '', '1.5', 'Estetyka, silikonowanie, rysy ', self.pt_id5],
               ['', '', '1.6', 'Kapturki srub NW, znaczki uziemień', self.pt_id6],
               ['', '', '1.7', 'Izolacja, uszczelki', self.pt_id7],
               ['', '', '1.8', 'Filtry', self.pt_id8],
               ['', '', '1.9', 'Wanna-szczelność-silikon-malowanie', self.pt_id9],
               ['2', 'Próby ruchowe', '2.1', self.equip_lista, self.pt_id10],
               ['', '', '2.2', 'Wentylatory-montaż-działanie', self.pt_id11],
               ['', '', '2.3', 'Czujniki temperatur-montaż-działanie', self.pt_id12],
               ['', '', '2.4', 'Montaż rozdzielnicy-wpięcie płytki ', self.pt_id13],
               ['', '', '2.5', 'Nagrzewnica-montaż-działanie-alarm', self.pt_id14],
               ['', '', '2.6', 'Drzwi-zamki-montaż-oznacznia-rączki', self.pt_id15],
               ['', '', '2.7', wydatek_info +' '+ self.pt_id16+' '+ unit_velocity,''],
               ['', '', '2.8', szczelnosc_info +' '+ self.pt_id17+' '+ unit_velocity ,leak_bool],
               ['3', 'Kompletacja', '3.1', 'Osłona rozdzielnicy-znak ostrzegawczy' , self.pt_id18],
               ['', '', '3.2', 'DTRka, karta produktu' , self.pt_id19],
               ['', '', '3.3', 'Oznaczenie KJ' , self.pt_id20],
               ['4', 'Uwagi',self.pt_id22,'', self.pt_id21],
               ['5', 'Nr zlecenia', str(self.pt_contens3), '' , ''],
               ['',  'Nr seryjny', str(self.pt_contens1), '' , ''],
               ['',  'Typ urzadzenia',str(self.pt_contens2), '' , ''],
               ['',  'Data badania KJ', self.pt_time, '' , ''],
               ['',  'Kontroler KJ', self.pt_id23, '' , ''],
               ['6',  'Ocena końcowa',self.summary_mark, '' , ''],


               ]



        t=Table(data,4*[1.3*inch], 30*[0.3*inch])
        t.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),

                               #('VALIGN',(0,0),(0,-1),'TOP'),
                               #('TEXTCOLOR',(0,0),(0,-1),colors.blue),
                               ('ALIGN',(0,0),(-1,-1),'CENTER'),
                               ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                               #('TEXTCOLOR',(0,-1),(-1,-1),colors.green),
                               ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
                               ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                               ('SPAN',(3,0),(4,0)),
                               ('SPAN',(2,1),(4,1)),
                               ('SPAN',(2,0),(4,0)),
                               ('SPAN',(0,0),(1,1)),
                               ('SPAN',(0,3),(0,11)),#
                               ('SPAN',(1,3),(1,11)),#
                               ('SPAN',(0,12),(0,19)),#
                               ('SPAN',(1,12),(1,19)),#
                               ('SPAN',(0,20),(0,22)),#
                               ('SPAN',(1,20),(1,22)),#
                               ('SPAN',(2,23),(3,23)),
                               ('SPAN',(2,24),(4,24)),
                               ('SPAN',(0,24),(0,28)),
                               ('SPAN',(2,25),(4,25)),
                               ('SPAN',(2,26),(4,26)),
                               ('SPAN',(2,27),(4,27)),
                               ('SPAN',(2,28),(4,28)),
                               ('SPAN',(2,29),(4,29)),
                               ('SPAN',(2,30),(4,30)),
                               ('FONTNAME',(0,0),(-1,-1),'DejaMono'),
                               ('FONTSIZE',(2,1),(2,4),12),
                               ('FONTSIZE',(2,0),(2,4),10),


                               ]))

        t._argW[0]=0.5*inch
        t._argW[1] = 1.5*inch
        t._argW[2] = 0.5*inch
        t._argW[3] = 3.2*inch
        t._argW[4] = 1.*inch
        t._argH[1]=0.4*inch
        t._argH[23]=0.5*inch
        t._argH[29]=0.5*inch
        elements.append(t)
        # write the document to disk
        doc.build(elements)
###############################################################################################################################################################################

    def filling_factory(self,State):

        pdfmetrics.registerFont(TTFont('DejaMono', 'DejaVuSansMono.ttf'))

        styleSheet = getSampleStyleSheet()

        style = styleSheet['BodyText']

        doc = SimpleDocTemplate("__file__"+"../../protokol/filling_protokol_KCX.pdf", pagesize=A4 )




        styleSheet = getSampleStyleSheet()

        style = styleSheet['BodyText']




        elements = []

        kom_NW=0






        if self.nagrz_dzial.get()=="Pozytyw":
            kom_NW="Sprawdzono w działaniu"

        if self.nagrz_dzial.get()=="Brak":
            kom_NW="BRAK"



        if  self.typ_ahu.get() in ('KCX 300 EPP', 'KCX 500 EPP', 'KCX 800 EPP', 'KCO 300 EPP', 'KCO 500 EPP', 'KCO 800 EPP' ):
           wydatek_liczba = 0
           szczelnosc_liczba = 0

        else :

            wydatek_liczba = float(self.ent_wydatek.get())
            szczelnosc_liczba = float(self.ent_nieszczel.get())



        kom_naw=0
        kom_szczel=0
        if self.typ_ahu.get()=='KCX-300' or self.typ_ahu.get()=='KCO-300' :
            kom_naw=str(round(wydatek_liczba*44,1))

        if self.typ_ahu.get()=='KCX-500' or self.typ_ahu.get()=='KCO-500':
            kom_naw=str(round(wydatek_liczba*72.34,1))

        if self.typ_ahu.get()=='KCX-800'or self.typ_ahu.get()=='KCO-800':
            kom_naw=str(round(wydatek_liczba*113.04,1))

        if self.typ_ahu.get()=='KCX-1200' or self.typ_ahu.get()=='KCO-1200':
            kom_naw=str(round(wydatek_liczba*176.6,1))



        if self.typ_ahu.get() == 'KCX-300' or self.typ_ahu.get() == 'KCO-300':
            kom_szczel = str(round(szczelnosc_liczba * 44, 2))

        if self.typ_ahu.get() == 'KCX-500' or self.typ_ahu.get() == 'KCO-500':
            kom_szczel = str(round(szczelnosc_liczba * 72.34, 2))

        if self.typ_ahu.get() == 'KCX-800' or self.typ_ahu.get() == 'KCO-800':
            kom_szczel = str(round(szczelnosc_liczba * 113.04, 2))

        if self.typ_ahu.get() == 'KCX-1200' or self.typ_ahu.get() == 'KCO-1200':
            kom_szczel = str(round(szczelnosc_liczba * 176.6, 2))



        naw_info_protoc = str(kom_naw) + '[m3/h]'
        szczel_info_protoc = str(szczelnosc_liczba) + '[m/s]' +' '+ str(kom_szczel) + '[m3/h]'



        if self.typ_ahu.get() in ('KCX 300 EPP', 'KCX 500 EPP', 'KCX 800 EPP','KCO 300 EPP', 'KCO 500 EPP', 'KCO 800 EPP'):
            naw_info_protoc = 'Nie dotyczy'
            szczel_info_protoc = 'Nie dotyczy'



        czas=str(self.now_d+ " "+self.now_h)

        data= [['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','',self.typ_ahu.get()+" \ "+ str(self.ent_nr_zlec.get()) ,'','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','',str(self.ent_nr_fab.get()),'','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','         '+self.equip_protocol_1,'','','','','','','',''],
               ['','      '+self.equip_protocol,'','','','','','','',''],
               ['','         '+self.equip_protocol_2,'','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','        '+str(kom_NW),'','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','   '+ naw_info_protoc,'','',''],
               ['','','','','','','','','',''],
               ['','','','','','   '+ szczel_info_protoc ,'','','',''],
               ['','','','','','','','','',''],
               ['','','',self.podpis.get(),'','','','',czas,''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ['','','','','','','','','',''],
               ]



        t=Table(data,10*[0.7*inch], 60*[0.1445*inch])
        t.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),

             ('VALIGN',(0,49),(9,49),'TOP'),
             ('ALIGN',(0,28),(9,28),'CENTER'),
             ('FONTNAME',(0,0),(-1,-1),'DejaMono'),
                               ]))


        elements.append(t)
        # write the document to disk
        doc.build(elements)
        self.protocol()
        self.equip_protocol = ''
        self.equip_protocol_1 = ''
        self.equip_protocol_2 = ''

    ##################################################################################################################################################################################

    def protocol_perm(self):

        self.protocol_permission=True

##################################################################################################################################################################################

    def protocol(self):

        if self.protocol_permission==True :

                row_protocol_KCX = open('__file__"+"../../protokol/raw_protocol_KCX.pdf', 'rb')
                pdfReader = PyPDF2.PdfFileReader(row_protocol_KCX)
                row_protocol_KCX_p0 = pdfReader.getPage(0)

                filling_protocol_KCX = (open("__file__"+"../../protokol/filling_protokol_KCX.pdf", 'rb'))
                pdfFillingReader = PyPDF2.PdfFileReader(filling_protocol_KCX)
                filling_protocol_KCX_p0 = pdfFillingReader.getPage(0)

                row_protocol_KCX_p0.mergePage(filling_protocol_KCX_p0)
                pdfWriter=PyPDF2.PdfFileWriter()
                pdfWriter.addPage(row_protocol_KCX_p0)

                resultPdfFile = open("__file__"+"../../protocols/"+self.pt_contens1+"_protocol_KCX.pdf", 'wb')
                pdfWriter.write(resultPdfFile)
                row_protocol_KCX .close()
                resultPdfFile.close()

                messagebox.showinfo('Wydruk protokołu', 'Protokół zostanie wydrukowany')
        else:
                messagebox.showinfo('Wydruk protokołu', 'Protokół nie zostanie wydrukowany')



##################################################################################################################################################################################

    def warning (self):
        messagebox.showwarning("Warning", "Popraw wpisana liczę")
        python = sys.executable
        os.execl(python, python, * sys.argv)
###################################################################################################################################################################################

    # funkcja przycisku archiwizacja - THE END
    def arch (self,n):
        State=[]


        cur.execute(
            """
            SELECT ID,nr_fabr,kod_prod, nr_zlec,identity,tab_znam_nakl,kla_en_oz_kr,zasl_raczki,est_silik_rysy,kapt_srub_uziem,\
            izol_uszcz,mon_filtr,tac_ociek,mon_bypas,went,czujn_temp,rozdz_mont,nagrz_dzial,drzwi,ent_wydatek,ent_nieszczel,\
            osl_rozdz,dtr_kart_prod,ozn_KJ,uwagi_bool,uwagi_txt,podpis,now_d,now_h FROM tab

            """)
        State_Train = cur.fetchall()
        State=State_Train[self.n:]


        self.grid_remove()


        app1 = Application(root,State,self.n)
        return State,app1,self.n

      #  except:

        #    messagebox.showinfo("Zapis danych", "Poza zakresem")

         #   python = sys.executable
         #   os.execl(python, python, * sys.argv)


# czesc glowna


root = Tk()
root.title("KJ KCX KCO")
root.geometry("660x770")



app = Application(root,State,n)



root.mainloop()

